import openpyxl


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.max_row


def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.max_column


def readData(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum, column=columnno).value


def writeData(file, sheetName, summary, result):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    getRowCount(file, sheetName)
    next_row = sheet.max_row + 1
    sheet.cell(row=next_row, column=1).value = next_row - 1
    sheet.cell(row=next_row, column=2).value = summary
    sheet.cell(row=next_row, column=3).value = result
    workbook.save(file)
